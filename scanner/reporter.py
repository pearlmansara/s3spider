"""
reporter.py - Write findings to an Excel (.xlsx) report.

Columns: Profile | Region | Bucket ARN | S3 File | Match Type | Matched Line
"""

import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger("s3spider")

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}

# Excel cell fill colors per severity
SEVERITY_FILLS = {
    "critical": "FFCCCC",   # light red
    "high":     "FFD9B3",   # light orange
    "medium":   "FFFF99",   # light yellow
    "low":      "E6E6E6",   # light grey
}


def write_excel(findings: list[dict], output_path: str):
    """
    Write all findings to an Excel workbook.

    findings: list of finding dicts produced by crawler.crawl_bucket()
    output_path: path to the .xlsx file to create/overwrite
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error(
            "openpyxl is required for Excel reporting. "
            "Install with: pip install openpyxl"
        )
        return

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, findings)

    # ── Findings sheet ────────────────────────────────────────────────────────
    ws_findings = wb.create_sheet("Findings")
    _write_findings(ws_findings, findings)

    # Save
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info(f"Report saved to {path}")


def _header_fill():
    try:
        from openpyxl.styles import PatternFill
        return PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    except Exception:
        return None


def _write_summary(ws, findings: list[dict]):
    """Write a summary sheet with counts by severity and by bucket."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return

    # Title row
    ws["A1"] = "S3Spider — Scan Summary"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:C1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:C2")

    # Severity counts
    row = 4
    ws.cell(row=row, column=1, value="Severity").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1

    from collections import Counter
    severity_counts = Counter(f["severity"] for f in findings)
    for sev in ["critical", "high", "medium", "low"]:
        count = severity_counts.get(sev, 0)
        cell_sev   = ws.cell(row=row, column=1, value=sev.upper())
        cell_count = ws.cell(row=row, column=2, value=count)
        fill_color = SEVERITY_FILLS.get(sev, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell_sev.fill   = fill
        cell_count.fill = fill
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Total Findings").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(findings)).font = Font(bold=True)

    # Bucket breakdown
    row += 2
    ws.cell(row=row, column=1, value="Bucket").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Profile").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Findings").font = Font(bold=True)
    row += 1

    bucket_counts: dict[tuple, int] = {}
    for f in findings:
        key = (f["bucket_arn"], f["profile"])
        bucket_counts[key] = bucket_counts.get(key, 0) + 1

    for (arn, profile), count in sorted(bucket_counts.items()):
        ws.cell(row=row, column=1, value=arn)
        ws.cell(row=row, column=2, value=profile)
        ws.cell(row=row, column=3, value=count)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12


def _write_findings(ws, findings: list[dict]):
    """Write the full findings detail sheet."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    headers = [
        "Profile",
        "Region",
        "Bucket ARN",
        "S3 File (Key)",
        "S3 URI",
        "Match Type",
        "Severity",
        "Line #",
        "Matched Line",
    ]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Sort findings by severity then bucket then key
    sorted_findings = sorted(
        findings,
        key=lambda x: (
            SEVERITY_ORDER.get(x["severity"], 99),
            x["bucket_name"],
            x["s3_key"],
            x["line_number"],
        ),
    )

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for row_idx, finding in enumerate(sorted_findings, start=2):
        sev         = finding["severity"]
        fill_color  = SEVERITY_FILLS.get(sev, "FFFFFF")
        row_fill    = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        values = [
            finding["profile"],
            finding["region"],
            finding["bucket_arn"],
            finding["s3_key"],
            finding["s3_uri"],
            finding["pattern_name"],
            sev.upper(),
            finding["line_number"],
            finding["line"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Apply severity color only to the Severity column (col 7)
            if col_idx == 7:
                cell.fill = row_fill
                cell.font = Font(bold=True)
            # Wrap matched line text
            if col_idx == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns (approximate)
    col_widths = [15, 15, 55, 55, 60, 30, 10, 8, 80]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-fit row heights for matched line column
    for row_idx in range(2, len(sorted_findings) + 2):
        ws.row_dimensions[row_idx].height = 40
